import logging
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import parse_cell_range, validate_cell_reference
from .exceptions import ValidationError

logger = logging.getLogger(__name__)

def validate_formula_in_cell_operation(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str
) -> dict[str, Any]:
    """Validate Excel formula before writing"""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")

        if not validate_cell_reference(cell):
            raise ValidationError(f"Invalid cell reference: {cell}")

        # First validate the provided formula's syntax
        is_valid, message = validate_formula(formula)
        if not is_valid:
            raise ValidationError(f"Invalid formula syntax: {message}")

        # Additional validation for cell references in formula
        cell_refs = re.findall(r'[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?', formula)
        for ref in cell_refs:
            if ':' in ref:  # Range reference
                start, end = ref.split(':')
                if not (validate_cell_reference(start) and validate_cell_reference(end)):
                    raise ValidationError(f"Invalid cell range reference in formula: {ref}")
            else:  # Single cell reference
                if not validate_cell_reference(ref):
                    raise ValidationError(f"Invalid cell reference in formula: {ref}")

        # Now check if there's a formula in the cell and compare
        sheet = wb[sheet_name]
        cell_obj = sheet[cell]
        current_formula = cell_obj.value

        # If cell has a formula (starts with =)
        if isinstance(current_formula, str) and current_formula.startswith('='):
            if formula.startswith('='):
                if current_formula != formula:
                    return {
                        "message": "Formula is valid but doesn't match cell content",
                        "valid": True,
                        "matches": False,
                        "cell": cell,
                        "provided_formula": formula,
                        "current_formula": current_formula
                    }
            else:
                if current_formula != f"={formula}":
                    return {
                        "message": "Formula is valid but doesn't match cell content",
                        "valid": True,
                        "matches": False,
                        "cell": cell,
                        "provided_formula": formula,
                        "current_formula": current_formula
                    }
                else:
                    return {
                        "message": "Formula is valid and matches cell content",
                        "valid": True,
                        "matches": True,
                        "cell": cell,
                        "formula": formula
                    }
        else:
            return {
                "message": "Formula is valid but cell contains no formula",
                "valid": True,
                "matches": False,
                "cell": cell,
                "provided_formula": formula,
                "current_content": str(current_formula) if current_formula else ""
            }

    except ValidationError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to validate formula: {e}")
        raise ValidationError(str(e))

def validate_range_in_sheet_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
) -> dict[str, Any]:
    """Validate if a range exists in a worksheet and return data range info."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Get actual data dimensions
        data_max_row = worksheet.max_row
        data_max_col = worksheet.max_column
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise ValidationError(f"Invalid range: {str(e)}")
            
        # If end not specified, use start
        if end_row is None:
            end_row = start_row
        if end_col is None:
            end_col = start_col
            
        # Validate bounds against maximum possible Excel limits
        is_valid, message = validate_range_bounds(
            worksheet, start_row, start_col, end_row, end_col
        )
        if not is_valid:
            raise ValidationError(message)
            
        range_str = f"{start_cell}" if end_cell is None else f"{start_cell}:{end_cell}"
        data_range_str = f"A1:{get_column_letter(data_max_col)}{data_max_row}"
        
        # Check if range is within data or extends beyond
        extends_beyond_data = (
            end_row > data_max_row or 
            end_col > data_max_col
        )
        
        return {
            "message": (
                f"Range '{range_str}' is valid. "
                f"Sheet contains data in range '{data_range_str}'"
            ),
            "valid": True,
            "range": range_str,
            "data_range": data_range_str,
            "extends_beyond_data": extends_beyond_data,
            "data_dimensions": {
                "max_row": data_max_row,
                "max_col": data_max_col,
                "max_col_letter": get_column_letter(data_max_col)
            }
        }
    except ValidationError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to validate range: {e}")
        raise ValidationError(str(e))

def validate_formula(formula: str) -> tuple[bool, str]:
    """Validate Excel formula syntax and safety"""
    if not formula.startswith("="):
        return False, "Formula must start with '='"

    # Remove the '=' prefix for validation
    formula = formula[1:]

    # Check for balanced parentheses
    parens = 0
    for c in formula:
        if c == "(":
            parens += 1
        elif c == ")":
            parens -= 1
        if parens < 0:
            return False, "Unmatched closing parenthesis"

    if parens > 0:
        return False, "Unclosed parenthesis"

    # Basic function name validation
    func_pattern = r"([A-Z]+)\("
    funcs = re.findall(func_pattern, formula)
    unsafe_funcs = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}

    for func in funcs:
        if func in unsafe_funcs:
            return False, f"Unsafe function: {func}"

    return True, "Formula is valid"


def validate_range_bounds(
    worksheet: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int | None = None,
    end_col: int | None = None,
) -> tuple[bool, str]:
    """Validate that cell range is within worksheet bounds"""
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    try:
        # Check start cell bounds
        if start_row < 1 or start_row > max_row:
            return False, f"Start row {start_row} out of bounds (1-{max_row})"
        if start_col < 1 or start_col > max_col:
            return False, (
                f"Start column {get_column_letter(start_col)} "
                f"out of bounds (A-{get_column_letter(max_col)})"
            )

        # If end cell specified, check its bounds
        if end_row is not None and end_col is not None:
            if end_row < start_row:
                return False, "End row cannot be before start row"
            if end_col < start_col:
                return False, "End column cannot be before start column"
            if end_row > max_row:
                return False, f"End row {end_row} out of bounds (1-{max_row})"
            if end_col > max_col:
                return False, (
                    f"End column {get_column_letter(end_col)} "
                    f"out of bounds (A-{get_column_letter(max_col)})"
                )

        return True, "Range is valid"
    except Exception as e:
        return False, f"Invalid range: {e!s}"