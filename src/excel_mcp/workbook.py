import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .exceptions import WorkbookError

logger = logging.getLogger(__name__)

def create_workbook(filepath: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name"""
    try:
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
            "workbook": wb
        }
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        raise WorkbookError(f"Failed to create workbook: {e!s}")

def get_or_create_workbook(filepath: str) -> Workbook:
    """Get existing workbook or create new one if it doesn't exist"""
    try:
        return load_workbook(filepath)
    except FileNotFoundError:
        return create_workbook(filepath)["workbook"]

def create_sheet(filepath: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        wb = load_workbook(filepath)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            raise WorkbookError(f"Sheet {sheet_name} already exists")

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(filepath)
        wb.close()
        return {"message": f"Sheet {sheet_name} created successfully"}
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise WorkbookError(str(e))

def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
            
        wb = load_workbook(filepath, read_only=True)
        
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime
        }
        
        if include_ranges:
            # Add used ranges for each sheet
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            info["used_ranges"] = ranges
            
        wb.close()
        return info
        
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook info: {e}")
        raise WorkbookError(str(e))
