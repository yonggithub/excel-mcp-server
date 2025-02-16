import logging
from typing import Any
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, PatternFill, Side

from .cell_utils import parse_cell_range
from .exceptions import SheetError, ValidationError

logger = logging.getLogger(__name__)

def copy_sheet(filepath: str, source_sheet: str, target_sheet: str) -> dict[str, Any]:
    """Copy a worksheet within the same workbook."""
    try:
        wb = load_workbook(filepath)
        if source_sheet not in wb.sheetnames:
            raise SheetError(f"Source sheet '{source_sheet}' not found")
            
        if target_sheet in wb.sheetnames:
            raise SheetError(f"Target sheet '{target_sheet}' already exists")
            
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = target_sheet
        
        wb.save(filepath)
        return {"message": f"Sheet '{source_sheet}' copied to '{target_sheet}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to copy sheet: {e}")
        raise SheetError(str(e))

def delete_sheet(filepath: str, sheet_name: str) -> dict[str, Any]:
    """Delete a worksheet from the workbook."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        if len(wb.sheetnames) == 1:
            raise SheetError("Cannot delete the only sheet in workbook")
            
        del wb[sheet_name]
        wb.save(filepath)
        return {"message": f"Sheet '{sheet_name}' deleted"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete sheet: {e}")
        raise SheetError(str(e))

def rename_sheet(filepath: str, old_name: str, new_name: str) -> dict[str, Any]:
    """Rename a worksheet."""
    try:
        wb = load_workbook(filepath)
        if old_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{old_name}' not found")
            
        if new_name in wb.sheetnames:
            raise SheetError(f"Sheet '{new_name}' already exists")
            
        sheet = wb[old_name]
        sheet.title = new_name
        wb.save(filepath)
        return {"message": f"Sheet renamed from '{old_name}' to '{new_name}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to rename sheet: {e}")
        raise SheetError(str(e))

def format_range_string(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    """Format range string from row and column indices."""
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

def copy_range(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_range: str,
    target_start: str | None = None,
) -> None:
    """Copy range from source worksheet to target worksheet."""
    # Parse source range
    if ':' in source_range:
        source_start, source_end = source_range.split(':')
    else:
        source_start = source_range
        source_end = None
        
    src_start_row, src_start_col, src_end_row, src_end_col = parse_cell_range(
        source_start, source_end
    )

    if src_end_row is None:
        src_end_row = src_start_row
        src_end_col = src_start_col

    if target_start is None:
        target_start = source_start

    tgt_start_row, tgt_start_col, _, _ = parse_cell_range(target_start)

    for i, row in enumerate(range(src_start_row, src_end_row + 1)):
        for j, col in enumerate(range(src_start_col, src_end_col + 1)):
            source_cell = source_ws.cell(row=row, column=col)
            target_cell = target_ws.cell(row=tgt_start_row + i, column=tgt_start_col + j)

            target_cell.value = source_cell.value

            try:
                # Copy font
                font_kwargs = {}
                if hasattr(source_cell.font, 'name'):
                    font_kwargs['name'] = source_cell.font.name
                if hasattr(source_cell.font, 'size'):
                    font_kwargs['size'] = source_cell.font.size
                if hasattr(source_cell.font, 'bold'):
                    font_kwargs['bold'] = source_cell.font.bold
                if hasattr(source_cell.font, 'italic'):
                    font_kwargs['italic'] = source_cell.font.italic
                if hasattr(source_cell.font, 'color'):
                    font_color = None
                    if source_cell.font.color:
                        font_color = source_cell.font.color.rgb
                    font_kwargs['color'] = font_color
                target_cell.font = Font(**font_kwargs)

                # Copy border
                new_border = Border()
                for side in ['left', 'right', 'top', 'bottom']:
                    source_side = getattr(source_cell.border, side)
                    if source_side and source_side.style:
                        side_color = source_side.color.rgb if source_side.color else None
                        setattr(new_border, side, Side(
                            style=source_side.style,
                            color=side_color
                        ))
                target_cell.border = new_border

                # Copy fill
                if hasattr(source_cell, 'fill'):
                    fill_kwargs = {'patternType': source_cell.fill.patternType}
                    if hasattr(source_cell.fill, 'fgColor') and source_cell.fill.fgColor:
                        fg_color = None
                        if hasattr(source_cell.fill.fgColor, 'rgb'):
                            fg_color = source_cell.fill.fgColor.rgb
                        fill_kwargs['fgColor'] = fg_color
                    if hasattr(source_cell.fill, 'bgColor') and source_cell.fill.bgColor:
                        bg_color = None
                        if hasattr(source_cell.fill.bgColor, 'rgb'):
                            bg_color = source_cell.fill.bgColor.rgb
                        fill_kwargs['bgColor'] = bg_color
                    target_cell.fill = PatternFill(**fill_kwargs)

                # Copy number format and alignment
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment

            except Exception:
                continue

def delete_range(worksheet: Worksheet, start_cell: str, end_cell: str | None = None) -> None:
    """Delete contents and formatting of a range."""
    start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

    if end_row is None:
        end_row = start_row
        end_col = start_col

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.number_format = "General"
            cell.alignment = None

def merge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    """Merge a range of cells."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

        if end_row is None or end_col is None:
            raise SheetError("Both start and end cells must be specified for merging")

        range_string = format_range_string(start_row, start_col, end_row, end_col)
        worksheet = wb[sheet_name]
        worksheet.merge_cells(range_string)
        wb.save(filepath)
        return {"message": f"Range '{range_string}' merged in sheet '{sheet_name}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to merge range: {e}")
        raise SheetError(str(e))

def unmerge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    """Unmerge a range of cells."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        
        if end_row is None or end_col is None:
            raise SheetError("Both start and end cells must be specified for unmerging")

        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        # Check if range is actually merged
        merged_ranges = worksheet.merged_cells.ranges
        target_range = range_string.upper()
        
        if not any(str(merged_range).upper() == target_range for merged_range in merged_ranges):
            raise SheetError(f"Range '{range_string}' is not merged")
            
        worksheet.unmerge_cells(range_string)
        wb.save(filepath)
        return {"message": f"Range '{range_string}' unmerged successfully"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to unmerge range: {e}")
        raise SheetError(str(e))

def copy_range_operation(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: str = None
) -> dict:
    """Copy a range of cells to another location."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found")
            raise ValidationError(f"Sheet '{sheet_name}' not found")

        source_ws = wb[sheet_name]
        target_ws = wb[target_sheet] if target_sheet else source_ws

        # Parse source range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(source_start, source_end)
        except ValueError as e:
            logger.error(f"Invalid source range: {e}")
            raise ValidationError(f"Invalid source range: {str(e)}")

        # Parse target starting point
        try:
            target_row = int(''.join(filter(str.isdigit, target_start)))
            target_col = column_index_from_string(''.join(filter(str.isalpha, target_start)))
        except ValueError as e:
            logger.error(f"Invalid target cell: {e}")
            raise ValidationError(f"Invalid target cell: {str(e)}")

        # Copy the range
        row_offset = target_row - start_row
        col_offset = target_col - start_col

        for i in range(start_row, end_row + 1):
            for j in range(start_col, end_col + 1):
                source_cell = source_ws.cell(row=i, column=j)
                target_cell = target_ws.cell(row=i + row_offset, column=j + col_offset)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)

        wb.save(filepath)
        return {"message": f"Range copied successfully"}

    except (ValidationError, SheetError):
        raise
    except Exception as e:
        logger.error(f"Failed to copy range: {e}")
        raise SheetError(f"Failed to copy range: {str(e)}")

def delete_range_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
    shift_direction: str = "up"
) -> dict[str, Any]:
    """Delete a range of cells and shift remaining cells."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            if end_row and end_row > worksheet.max_row:
                raise SheetError(f"End row {end_row} out of bounds (1-{worksheet.max_row})")
            if end_col and end_col > worksheet.max_column:
                raise SheetError(f"End column {end_col} out of bounds (1-{worksheet.max_column})")
        except ValueError as e:
            raise SheetError(f"Invalid range: {str(e)}")
            
        # Validate shift direction
        if shift_direction not in ["up", "left"]:
            raise ValidationError(f"Invalid shift direction: {shift_direction}. Must be 'up' or 'left'")
            
        range_string = format_range_string(
            start_row, start_col,
            end_row or start_row,
            end_col or start_col
        )
        
        # Delete range contents
        delete_range(worksheet, start_cell, end_cell)
        
        # Shift cells if needed
        if shift_direction == "up":
            worksheet.delete_rows(start_row, (end_row or start_row) - start_row + 1)
        elif shift_direction == "left":
            worksheet.delete_cols(start_col, (end_col or start_col) - start_col + 1)
            
        wb.save(filepath)
        
        return {"message": f"Range {range_string} deleted successfully"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete range: {e}")
        raise SheetError(str(e))
