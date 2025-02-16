import logging
from typing import Any, Dict

from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Protection, Font,
    Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    FormulaRule, CellIsRule
)

from .workbook import get_or_create_workbook
from .cell_utils import parse_cell_range, validate_cell_reference
from .exceptions import ValidationError, FormattingError

logger = logging.getLogger(__name__)

def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: int = None,
    font_color: str = None,
    bg_color: str = None,
    border_style: str = None,
    border_color: str = None,
    number_format: str = None,
    alignment: str = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Dict[str, Any] = None,
    conditional_format: Dict[str, Any] = None
) -> Dict[str, Any]:
    """Apply formatting to a range of cells.
    
    This function handles all Excel formatting operations including:
    - Font properties (bold, italic, size, color, etc.)
    - Cell fill/background color
    - Borders (style and color)
    - Number formatting
    - Alignment and text wrapping
    - Cell merging
    - Protection
    - Conditional formatting
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell reference
        end_cell: Optional ending cell reference
        bold: Whether to make text bold
        italic: Whether to make text italic
        underline: Whether to underline text
        font_size: Font size in points
        font_color: Font color (hex code)
        bg_color: Background color (hex code)
        border_style: Border style (thin, medium, thick, double)
        border_color: Border color (hex code)
        number_format: Excel number format string
        alignment: Text alignment (left, center, right, justify)
        wrap_text: Whether to wrap text
        merge_cells: Whether to merge the range
        protection: Cell protection settings
        conditional_format: Conditional formatting rules
        
    Returns:
        Dictionary with operation status
    """
    try:
        # Validate cell references
        if not validate_cell_reference(start_cell):
            raise ValidationError(f"Invalid start cell reference: {start_cell}")
            
        if end_cell and not validate_cell_reference(end_cell):
            raise ValidationError(f"Invalid end cell reference: {end_cell}")
            
        wb = get_or_create_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
            
        sheet = wb[sheet_name]
        
        # Get cell range coordinates
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise ValidationError(f"Invalid cell range: {str(e)}")
        
        # If no end cell specified, use start cell coordinates
        if end_row is None:
            end_row = start_row
        if end_col is None:
            end_col = start_col
            
        # Apply font formatting
        font_args = {
            "bold": bold,
            "italic": italic,
            "underline": 'single' if underline else None,
        }
        if font_size is not None:
            font_args["size"] = font_size
        if font_color is not None:
            try:
                # Ensure color has FF prefix for full opacity
                font_color = font_color if font_color.startswith('FF') else f'FF{font_color}'
                font_args["color"] = Color(rgb=font_color)
            except ValueError as e:
                raise FormattingError(f"Invalid font color: {str(e)}")
        font = Font(**font_args)
        
        # Apply fill
        fill = None
        if bg_color is not None:
            try:
                # Ensure color has FF prefix for full opacity
                bg_color = bg_color if bg_color.startswith('FF') else f'FF{bg_color}'
                fill = PatternFill(
                    start_color=Color(rgb=bg_color),
                    end_color=Color(rgb=bg_color),
                    fill_type='solid'
                )
            except ValueError as e:
                raise FormattingError(f"Invalid background color: {str(e)}")
        
        # Apply borders
        border = None
        if border_style is not None:
            try:
                border_color = border_color if border_color else "000000"
                border_color = border_color if border_color.startswith('FF') else f'FF{border_color}'
                side = Side(
                    style=border_style,
                    color=Color(rgb=border_color)
                )
                border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side
                )
            except ValueError as e:
                raise FormattingError(f"Invalid border settings: {str(e)}")
            
        # Apply alignment
        align = None
        if alignment is not None or wrap_text:
            try:
                align = Alignment(
                    horizontal=alignment,
                    vertical='center',
                    wrap_text=wrap_text
                )
            except ValueError as e:
                raise FormattingError(f"Invalid alignment settings: {str(e)}")
            
        # Apply protection
        protect = None
        if protection is not None:
            try:
                protect = Protection(**protection)
            except ValueError as e:
                raise FormattingError(f"Invalid protection settings: {str(e)}")
            
        # Apply formatting to range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = font
                if fill is not None:
                    cell.fill = fill
                if border is not None:
                    cell.border = border
                if align is not None:
                    cell.alignment = align
                if protect is not None:
                    cell.protection = protect
                if number_format is not None:
                    cell.number_format = number_format
                    
        # Merge cells if requested
        if merge_cells and end_cell:
            try:
                range_str = f"{start_cell}:{end_cell}"
                sheet.merge_cells(range_str)
            except ValueError as e:
                raise FormattingError(f"Failed to merge cells: {str(e)}")
            
        # Apply conditional formatting
        if conditional_format is not None:
            range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
            rule_type = conditional_format.get('type')
            if not rule_type:
                raise FormattingError("Conditional format type not specified")
                
            params = conditional_format.get('params', {})
            
            # Handle fill parameter for cell_is rule
            if rule_type == 'cell_is' and 'fill' in params:
                fill_params = params['fill']
                if isinstance(fill_params, dict):
                    try:
                        fill_color = fill_params.get('fgColor', 'FFC7CE')  # Default to light red
                        fill_color = fill_color if fill_color.startswith('FF') else f'FF{fill_color}'
                        params['fill'] = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type='solid'
                        )
                    except ValueError as e:
                        raise FormattingError(f"Invalid conditional format fill color: {str(e)}")
            
            try:
                if rule_type == 'color_scale':
                    rule = ColorScaleRule(**params)
                elif rule_type == 'data_bar':
                    rule = DataBarRule(**params)
                elif rule_type == 'icon_set':
                    rule = IconSetRule(**params)
                elif rule_type == 'formula':
                    rule = FormulaRule(**params)
                elif rule_type == 'cell_is':
                    rule = CellIsRule(**params)
                else:
                    raise FormattingError(f"Invalid conditional format type: {rule_type}")
                    
                sheet.conditional_formatting.add(range_str, rule)
            except Exception as e:
                raise FormattingError(f"Failed to apply conditional formatting: {str(e)}")
            
        wb.save(filepath)
        
        range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
        return {
            "message": f"Applied formatting to range {range_str}",
            "range": range_str
        }
        
    except (ValidationError, FormattingError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply formatting: {e}")
        raise FormattingError(str(e))
