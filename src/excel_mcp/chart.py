from typing import Any, Optional, Dict
import logging
from enum import Enum

from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, ScatterChart, 
    AreaChart, Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker, OneCellAnchor, SpreadsheetDrawing
)
from openpyxl.utils import column_index_from_string

from .cell_utils import parse_cell_range
from .exceptions import ValidationError, ChartError

logger = logging.getLogger(__name__)

class ChartType(str, Enum):
    """Supported chart types"""
    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"
    BUBBLE = "bubble"
    STOCK = "stock"
    SURFACE = "surface"
    RADAR = "radar"

class ChartStyle:
    """Chart style configuration"""
    def __init__(
        self,
        title_size: int = 14,
        title_bold: bool = True,
        axis_label_size: int = 12,
        show_legend: bool = True,
        legend_position: str = "r",
        show_data_labels: bool = True,
        grid_lines: bool = False,
        style_id: int = 2
    ):
        self.title_size = title_size
        self.title_bold = title_bold
        self.axis_label_size = axis_label_size
        self.show_legend = show_legend
        self.legend_position = legend_position
        self.show_data_labels = show_data_labels
        self.grid_lines = grid_lines
        self.style_id = style_id

def create_chart_in_sheet(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    style: Optional[Dict] = None
) -> dict[str, Any]:
    """Create chart in sheet with enhanced styling options"""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found")
            raise ValidationError(f"Sheet '{sheet_name}' not found")

        worksheet = wb[sheet_name]

        # Initialize collections if they don't exist
        if not hasattr(worksheet, '_drawings'):
            worksheet._drawings = []
        if not hasattr(worksheet, '_charts'):
            worksheet._charts = []

        # Parse the data range
        if "!" in data_range:
            range_sheet_name, cell_range = data_range.split("!")
            if range_sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{range_sheet_name}' referenced in data range not found")
                raise ValidationError(f"Sheet '{range_sheet_name}' referenced in data range not found")
        else:
            cell_range = data_range

        try:
            start_cell, end_cell = cell_range.split(":")
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            logger.error(f"Invalid data range format: {e}")
            raise ValidationError(f"Invalid data range format: {str(e)}")

        # Validate chart type
        chart_classes = {
            "line": LineChart,
            "bar": BarChart,
            "pie": PieChart,
            "scatter": ScatterChart,
            "area": AreaChart
        }
        
        chart_type_lower = chart_type.lower()
        ChartClass = chart_classes.get(chart_type_lower)
        if not ChartClass:
            logger.error(f"Unsupported chart type: {chart_type}")
            raise ValidationError(
                f"Unsupported chart type: {chart_type}. "
                f"Supported types: {', '.join(chart_classes.keys())}"
            )
            
        chart = ChartClass()
        
        # Basic chart settings
        chart.title = title
        if hasattr(chart, "x_axis"):
            chart.x_axis.title = x_axis
        if hasattr(chart, "y_axis"):
            chart.y_axis.title = y_axis

        try:
            # Create data references
            if chart_type_lower == "scatter":
                # For scatter charts, create series for each pair of columns
                for col in range(start_col + 1, end_col + 1):
                    x_values = Reference(
                        worksheet,
                        min_row=start_row + 1,
                        max_row=end_row,
                        min_col=start_col
                    )
                    y_values = Reference(
                        worksheet,
                        min_row=start_row + 1,
                        max_row=end_row,
                        min_col=col
                    )
                    series = Series(y_values, x_values, title_from_data=True)
                    chart.series.append(series)
            else:
                # For other chart types
                data = Reference(
                    worksheet,
                    min_row=start_row,
                    max_row=end_row,
                    min_col=start_col + 1,
                    max_col=end_col
                )
                cats = Reference(
                    worksheet,
                    min_row=start_row + 1,
                    max_row=end_row,
                    min_col=start_col
                )
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
        except Exception as e:
            logger.error(f"Failed to create chart data references: {e}")
            raise ChartError(f"Failed to create chart data references: {str(e)}")

        # Apply style if provided
        try:
            if style:
                if style.get("show_legend", True):
                    chart.legend = Legend()
                    chart.legend.position = style.get("legend_position", "r")
                else:
                    chart.legend = None

                if style.get("show_data_labels", False):
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showVal = True

                if style.get("grid_lines", False):
                    if hasattr(chart, "x_axis"):
                        chart.x_axis.majorGridlines = ChartLines()
                    if hasattr(chart, "y_axis"):
                        chart.y_axis.majorGridlines = ChartLines()
        except Exception as e:
            logger.error(f"Failed to apply chart style: {e}")
            raise ChartError(f"Failed to apply chart style: {str(e)}")

        # Set chart size
        chart.width = 15
        chart.height = 7.5

        # Create drawing and anchor
        try:
            drawing = SpreadsheetDrawing()
            drawing.chart = chart

            # Validate target cell format
            if not target_cell or not any(c.isalpha() for c in target_cell) or not any(c.isdigit() for c in target_cell):
                raise ValidationError(f"Invalid target cell format: {target_cell}")

            # Create anchor
            col = column_index_from_string(target_cell[0]) - 1
            row = int(target_cell[1:]) - 1
            anchor = OneCellAnchor()
            anchor._from = AnchorMarker(col=col, row=row)
            drawing.anchor = anchor

            # Add to worksheet
            worksheet._drawings.append(drawing)
            worksheet._charts.append(chart)
        except ValueError as e:
            logger.error(f"Invalid target cell: {e}")
            raise ValidationError(f"Invalid target cell: {str(e)}")
        except Exception as e:
            logger.error(f"Failed to create chart drawing: {e}")
            raise ChartError(f"Failed to create chart drawing: {str(e)}")

        try:
            wb.save(filepath)
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            raise ChartError(f"Failed to save workbook with chart: {str(e)}")

        return {
            "message": f"{chart_type.capitalize()} chart created successfully",
            "details": {
                "type": chart_type,
                "location": target_cell,
                "data_range": data_range
            }
        }
        
    except (ValidationError, ChartError):
        raise
    except Exception as e:
        logger.error(f"Unexpected error creating chart: {e}")
        raise ChartError(f"Unexpected error creating chart: {str(e)}")
